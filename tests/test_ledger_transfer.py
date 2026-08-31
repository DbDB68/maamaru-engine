import asyncio
import csv
import io
import json
import sqlite3
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from touken.ledger_transfer import (
    apply_import_preview,
    create_import_preview,
    export_ledger_csv,
    export_ledger_xlsx,
)
from touken.telemetry import TelemetryStore


class LedgerTransferTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.store = TelemetryStore(self.root / "telemetry.db")
        self.now = time.time() - 3600

    def tearDown(self):
        self.store.close()
        self.temp.cleanup()

    def _event(self, ts, event_type, payload, run_id="run-1", script="daily"):
        self.store._conn().execute(
            "INSERT INTO events(ts, run_id, script, event_type, payload) "
            "VALUES (?, ?, ?, ?, ?)",
            (ts, run_id, script, event_type, json.dumps(payload, ensure_ascii=False)),
        )
        self.store._conn().commit()

    def test_xlsx_exports_player_sheets_and_roundtrips_only_manual_rows(self):
        self._event(self.now, "inventory.captured",
                    {"phase": "before", "resources": {"小判": 1000}})
        self._event(self.now + 120, "resource.change",
                    {"resource": "小判", "delta": 100, "source": "task_rewards",
                     "note": "日课奖励"})
        self._event(self.now + 240, "inventory.captured",
                    {"phase": "after", "resources": {"小判": 1100}})
        self.store.add_human_report(
            occurred_at=self.now + 300, activities=["买东西"], note="自己花的",
            resource="小判", claimed_delta=-50,
        )
        self.store.add_human_report_group(
            occurred_at=self.now + 330, activities=["收件箱"], note="一起领的",
            entries={"木炭": 25, "玉钢": 30},
        )
        self.store.add_manual_inventory({"小判": 1050, "木炭": 500}, self.now + 360)
        self.store.add_manual_session(
            script="osaka", started_at=self.now + 400, ended_at=self.now + 1000,
            loops=3, note="自己打的",
        )

        data = export_ledger_xlsx(self.store)
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False)

        self.assertEqual(
            workbook.sheetnames,
            ["使用说明", "完整流水", "当前家底", "每日汇总", "可再次导入"],
        )
        full_rows = list(workbook["完整流水"].iter_rows(values_only=True))
        self.assertTrue(any(row[1] == "まあ丸自动" and row[4] == "小判" for row in full_rows[1:]))
        self.assertFalse(any(row[1] == "未归因" and row[4] == "小判" and row[5] == -50
                             for row in full_rows[1:]))
        import_rows = list(workbook["可再次导入"].iter_rows(values_only=True))
        self.assertTrue(import_rows)
        self.assertTrue(all(row[-1] == "你手动" for row in import_rows[1:]))
        workbook.close()

        same_store_preview = create_import_preview(self.store, data, "ledger.xlsx")
        self.assertEqual(same_store_preview["counts"]["new"], 0)
        self.assertGreaterEqual(same_store_preview["counts"]["duplicate"], 4)

        other_store = TelemetryStore(self.root / "other.db")
        try:
            other_preview = create_import_preview(other_store, data, "ledger.xlsx")
            self.assertGreaterEqual(other_preview["counts"]["new"], 4)
            self.assertEqual(other_preview["counts"]["conflict"], 0)
            result = apply_import_preview(
                other_store, other_preview["preview_id"], self.root / "backups")
            self.assertEqual(result["imported"], other_preview["counts"]["new"])
            imported_inventory = other_store.manual_inventory()
            self.assertEqual(len(imported_inventory), 1)
            self.assertEqual(imported_inventory[0]["resources"], {"小判": 1050, "木炭": 500})
            imported_groups = other_store._conn().execute(
                "SELECT group_id, COUNT(*) AS total FROM human_reports "
                "WHERE group_id IS NOT NULL GROUP BY group_id").fetchall()
            self.assertEqual([(row["total"]) for row in imported_groups], [2])
        finally:
            other_store.close()

    def test_csv_export_has_utf8_bom_and_complete_ledger_header(self):
        self._event(self.now, "inventory.captured", {"resources": {"木炭": 10}})
        self._event(self.now + 60, "resource.change",
                    {"resource": "木炭", "delta": 5, "source": "daily", "note": "日课"})
        self._event(self.now + 120, "inventory.captured", {"resources": {"木炭": 15}})

        data = export_ledger_csv(self.store)

        self.assertTrue(data.startswith(b"\xef\xbb\xbf"))
        rows = list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))
        self.assertEqual(rows[0][:6], ["时间", "来源", "类型", "玩法", "资源", "变化"])
        self.assertTrue(any(row[4] == "木炭" and row[5] == "5" for row in rows[1:]))

    def test_exports_escape_spreadsheet_formula_text(self):
        self.store.add_human_report(
            occurred_at=self.now, activities=[], note="=HYPERLINK(\"https://invalid.example\")",
            resource="砥石", claimed_delta=1,
        )

        workbook = load_workbook(
            io.BytesIO(export_ledger_xlsx(self.store)), read_only=True, data_only=False)
        full_rows = list(workbook["完整流水"].iter_rows())
        note_cell = next(row[6] for row in full_rows[1:] if row[1].value == "你手动")
        self.assertEqual(note_cell.data_type, "s")
        self.assertTrue(note_cell.value.startswith("'="))
        workbook.close()
        csv_text = export_ledger_csv(self.store).decode("utf-8-sig")
        self.assertIn("'=HYPERLINK", csv_text)

    def test_import_requires_conflict_confirmation_and_backs_up_before_write(self):
        self._event(self.now, "inventory.captured", {"resources": {"小判": 1000}})
        source = io.StringIO(newline="")
        writer = csv.writer(source)
        writer.writerow(["记录类型", "时间", "资源", "数额", "备注", "来源"])
        writer.writerow(["家底", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(self.now)),
                         "小判", 900, "旧表", "你手动"])
        writer.writerow(["收支", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(self.now + 60)),
                         "木炭", -20, "旧表", "你手动"])
        data = source.getvalue().encode("utf-8-sig")

        preview = create_import_preview(self.store, data, "old.csv")
        self.assertEqual(preview["counts"]["conflict"], 1)
        self.assertEqual(preview["counts"]["new"], 1)
        with self.assertRaisesRegex(ValueError, "冲突"):
            apply_import_preview(self.store, preview["preview_id"], self.root / "backups")

        before_event_count = self.store._conn().execute("SELECT COUNT(*) FROM events").fetchone()[0]
        result = apply_import_preview(
            self.store, preview["preview_id"], self.root / "backups", accept_conflicts=True)

        self.assertEqual(result["imported"], 2)
        backup = Path(result["backup"])
        self.assertTrue((backup / "telemetry.db").is_file())
        self.assertTrue((backup / "manifest.json").is_file())
        backup_conn = sqlite3.connect(backup / "telemetry.db")
        try:
            self.assertEqual(backup_conn.execute("SELECT COUNT(*) FROM events").fetchone()[0],
                             before_event_count)
            self.assertEqual(backup_conn.execute("SELECT COUNT(*) FROM human_reports").fetchone()[0], 0)
        finally:
            backup_conn.close()
        self.assertEqual(self.store._conn().execute("SELECT COUNT(*) FROM events").fetchone()[0],
                         before_event_count + 1)
        self.assertEqual(self.store._conn().execute("SELECT COUNT(*) FROM human_reports").fetchone()[0], 1)
        imported = self.store.manual_inventory()[0]
        self.assertEqual(imported["source"], "manual_import")
        self.assertEqual(imported["resources"], {"小判": 900})

    def test_all_duplicates_make_no_backup_and_no_write(self):
        self.store.add_human_report(
            occurred_at=self.now, activities=[], note="已有", resource="玉钢", claimed_delta=10)
        source = ("记录类型,时间,资源,数额,备注,来源\n"
                  f"收支,{time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(self.now))},玉钢,10,已有,你手动\n")
        preview = create_import_preview(self.store, source.encode("utf-8-sig"), "old.csv")

        result = apply_import_preview(self.store, preview["preview_id"], self.root / "backups")

        self.assertEqual(result["imported"], 0)
        self.assertIsNone(result["backup"])
        self.assertFalse((self.root / "backups").exists())

    def test_public_api_exports_previews_and_applies(self):
        from panel.server import (api_ledger_export, api_ledger_import_apply,
                                  api_ledger_import_preview)

        class _Req:
            def __init__(self, *, raw=b"", payload=None):
                self.raw = raw
                self.payload = payload

            async def body(self):
                return self.raw

            async def json(self):
                return self.payload

        source = ("记录类型,时间,资源,数额,备注,来源\n"
                  f"收支,{time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(self.now))},"
                  "冷却材,12,旧账,你手动\n").encode("utf-8-sig")
        backup_root = self.root / "api-backups"
        with patch("touken.telemetry._store", self.store), \
                patch("panel.server.BACKUP_DIR", backup_root):
            exported = asyncio.run(api_ledger_export(format="xlsx"))
            preview = asyncio.run(api_ledger_import_preview(
                _Req(raw=source), filename="old.csv"))
            applied = asyncio.run(api_ledger_import_apply(_Req(payload={
                "preview_id": preview["preview_id"], "accept_conflicts": False,
            })))

        self.assertTrue(exported.body.startswith(b"PK"))
        self.assertIn("maamaru-ledger-", exported.headers["content-disposition"])
        self.assertEqual(preview["counts"]["new"], 1)
        self.assertEqual(applied["imported"], 1)
        self.assertTrue(Path(applied["backup"]).is_dir())


if __name__ == "__main__":
    unittest.main()
