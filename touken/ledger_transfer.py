"""本丸账房的 Excel/CSV 导入导出。

自动流水只读导出；导入只能创建用户手动记录，不会改写 runs 或
任何脚本自动确认的 events。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import secrets
import sqlite3
import threading
import time
from collections import defaultdict
from datetime import date, datetime, time as datetime_time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .telemetry import LEDGER_RESOURCES, TelemetryStore, _LEDGER_TZ, _json, _loads


TRANSFER_SCHEMA_VERSION = 1
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
PREVIEW_TTL_SECONDS = 30 * 60
_IMPORT_SHEET = "可再次导入"
_RESOURCE_SET = frozenset(LEDGER_RESOURCES)
_PREVIEW_LOCK = threading.Lock()
_PREVIEWS: dict[str, dict] = {}

_MANUAL_SCRIPT_LABELS = {
    "osaka": "大阪城", "raid": "联队战", "edocastle": "江户城",
    "sortie": "合战场", "yosari": "异去", "pumpkin": "季节活动",
}
_SCRIPT_LABELS = {
    **_MANUAL_SCRIPT_LABELS,
    "daily": "一键日课", "sakura": "刷花", "practice": "演练",
    "expedition": "远征", "dispatch": "派遣远征", "forge": "锻刀",
    "repair": "手入", "sugar": "炼糖", "inbox_supplies": "收杂物箱",
    "snapshot": "库存快照",
}
_SCRIPT_KEYS = {**{key: key for key in _MANUAL_SCRIPT_LABELS},
                **{label: key for key, label in _MANUAL_SCRIPT_LABELS.items()}}
_CONFIDENCE_LABELS = {"confirmed": "已确认", "high": "高", "medium": "中", "low": "低"}
_HEADER_ALIASES = {
    "kind": {"记录类型", "类型", "kind", "type"},
    "occurred_at": {"时间", "日期", "记录时间", "开始时间", "occurred_at", "time", "date"},
    "ended_at": {"结束时间", "ended_at", "end_time"},
    "resource": {"资源", "资源名", "resource"},
    "amount": {"数额", "数量", "收支变化", "变化", "delta", "amount", "value"},
    "script": {"玩法", "活动", "script"},
    "loops": {"圈数", "次数", "loops", "count"},
    "note": {"备注", "说明", "note"},
    "activities": {"标签", "操作", "activities", "tags"},
    "group": {"组号", "group", "group_id"},
    "origin": {"来源", "origin", "source"},
}


def _timestamp(value: Any) -> float:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime_time.min)
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if math.isfinite(number) and number > 1_000_000_000:
            return number
        raise ValueError("时间不正确")
    else:
        text = str(value or "").strip()
        if not text:
            raise ValueError("没有填时间")
        text = text.replace("年", "-").replace("月", "-").replace("日", " ")
        parsed = None
        for fmt in (
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
        ):
            try:
                parsed = datetime.strptime(text.strip(), fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            try:
                parsed = datetime.fromisoformat(text)
            except ValueError as exc:
                raise ValueError(f"不认识的时间：{text}") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=_LEDGER_TZ)
    return parsed.timestamp()


def _display_time(ts: float | None) -> str:
    if ts is None:
        return ""
    return datetime.fromtimestamp(float(ts), _LEDGER_TZ).strftime("%Y-%m-%d %H:%M:%S")


def _confidence_label(value: Any) -> str:
    text = str(value or "").strip()
    return _CONFIDENCE_LABELS.get(text, text)


def _safe_cell(value: Any) -> Any:
    """阻止手动备注在 Excel/CSV 中被当成公式执行。"""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + value
    return value


def _clean_int(value: Any, label: str, *, positive: bool = False) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label}要填整数")
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        raise ValueError(f"{label}要填整数") from None
    if not math.isfinite(number) or not number.is_integer():
        raise ValueError(f"{label}要填整数")
    result = int(number)
    if positive and result <= 0:
        raise ValueError(f"{label}要大于 0")
    return result


def _canonical_headers(values: list[Any]) -> dict[str, int]:
    normalized = {str(value or "").strip().lower(): index for index, value in enumerate(values)}
    result = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias.lower() in normalized:
                result[canonical] = normalized[alias.lower()]
                break
    return result


def _cell(row: list[Any], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    return row[index] if index is not None and index < len(row) else None


def _normalize_row(row: list[Any], headers: dict[str, int], row_number: int) -> dict | None:
    if not any(value not in (None, "") for value in row):
        return None
    origin = str(_cell(row, headers, "origin") or "").strip().lower()
    if origin and any(mark in origin for mark in ("まあ丸", "自动", "maamaru", "未归因")):
        return {"row": row_number, "ignored": True,
                "reason": "这是まあ丸自动流水，只导出不回写"}
    kind_text = str(_cell(row, headers, "kind") or "").strip().lower()
    if any(mark in kind_text for mark in ("家底", "库存", "inventory")):
        kind = "inventory"
    elif any(mark in kind_text for mark in ("活动", "挂机", "session")):
        kind = "session"
    else:
        kind = "transaction"
    occurred_at = _timestamp(_cell(row, headers, "occurred_at"))
    if occurred_at > time.time() + 300:
        raise ValueError("时间在未来")
    note = str(_cell(row, headers, "note") or "").strip()[:300]
    activities = [item.strip()[:40] for item in
                  str(_cell(row, headers, "activities") or "").replace("，", ",").split(",")
                  if item.strip()][:20]
    if kind in {"transaction", "inventory"}:
        resource = str(_cell(row, headers, "resource") or "").strip()
        if resource not in _RESOURCE_SET:
            raise ValueError(f"不认识的资源：{resource or '空'}")
        amount = _clean_int(_cell(row, headers, "amount"), "数额")
        if kind == "inventory" and amount < 0:
            raise ValueError("家底不能是负数")
        if kind == "transaction" and not amount:
            raise ValueError("收支变化不能是 0")
        return {"kind": kind, "row": row_number, "occurred_at": occurred_at,
                "resource": resource, "amount": amount, "note": note,
                "activities": activities,
                "group_key": str(_cell(row, headers, "group") or "").strip()[:80]}
    script_value = str(_cell(row, headers, "script") or "").strip()
    script = _SCRIPT_KEYS.get(script_value)
    if not script:
        raise ValueError(f"不认识的玩法：{script_value or '空'}")
    ended_at = _timestamp(_cell(row, headers, "ended_at"))
    loops = _clean_int(_cell(row, headers, "loops"), "圈数", positive=True)
    if ended_at <= occurred_at:
        raise ValueError("结束时间要晚于开始时间")
    return {"kind": kind, "row": row_number, "occurred_at": occurred_at,
            "ended_at": ended_at, "script": script, "loops": loops, "note": note}


def _parse_rows(data: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    if not data:
        raise ValueError("选中的文件是空的")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("文件超过 10 MB，请先拆分再导入")
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        text = data.decode("utf-8-sig")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect))
    elif suffix == ".xlsx":
        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("这个 Excel 文件无法读取") from exc
        sheet = workbook[_IMPORT_SHEET] if _IMPORT_SHEET in workbook.sheetnames else workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        raise ValueError("只支持 .xlsx 和 .csv")
    if not rows:
        raise ValueError("表格里没有数据")
    headers = _canonical_headers(rows[0])
    if "occurred_at" not in headers or "amount" not in headers:
        raise ValueError("没找到“时间”和“数额”列，请使用まあ丸导出表的“可再次导入”格式")
    entries, issues = [], []
    for row_number, row in enumerate(rows[1:MAX_IMPORT_ROWS + 1], start=2):
        try:
            item = _normalize_row(row, headers, row_number)
            if item and item.get("ignored"):
                issues.append(item)
            elif item:
                entries.append(item)
        except ValueError as exc:
            issues.append({"row": row_number, "ignored": False, "reason": str(exc)})
    if len(rows) > MAX_IMPORT_ROWS + 1:
        issues.append({"row": MAX_IMPORT_ROWS + 2, "ignored": False,
                       "reason": f"只预览前 {MAX_IMPORT_ROWS} 行，请拆分文件"})
    return entries, issues


def _analyse_entries(store: TelemetryStore, entries: list[dict]) -> list[dict]:
    conn = store._conn()
    analysed = []
    for entry in entries:
        status, detail = "new", "可导入"
        if entry["kind"] == "transaction":
            rows = conn.execute(
                "SELECT id, claimed_delta, note FROM human_reports "
                "WHERE source = 'proactive' AND resource = ? AND ABS(occurred_at - ?) <= 1",
                (entry["resource"], entry["occurred_at"])).fetchall()
            if rows:
                exact = any(int(row["claimed_delta"] or 0) == entry["amount"]
                            and str(row["note"] or "").strip() == entry["note"] for row in rows)
                status, detail = (("duplicate", "同一笔手动收支已存在") if exact else
                                  ("conflict", "同一时间的同种资源已有不同收支"))
        elif entry["kind"] == "inventory":
            rows = conn.execute(
                "SELECT id, run_id, script, payload FROM events "
                "WHERE event_type = 'inventory.captured' AND ABS(ts - ?) <= 5",
                (entry["occurred_at"],)).fetchall()
            for row in rows:
                resources = _loads(row["payload"], {}).get("resources") or {}
                if entry["resource"] not in resources:
                    continue
                same = int(resources[entry["resource"]]) == entry["amount"]
                if same:
                    status, detail = "duplicate", "同一时间的家底读数已存在"
                else:
                    owner = "まあ丸自动盘点" if row["run_id"] or row["script"] != "manual" else "手动家底"
                    status, detail = "conflict", f"与同时的{owner}数量不同"
                break
        else:
            rows = conn.execute(
                "SELECT id, ended_at, loops, note FROM manual_sessions "
                "WHERE script = ? AND ABS(started_at - ?) <= 1",
                (entry["script"], entry["occurred_at"])).fetchall()
            if rows:
                exact = any(abs(float(row["ended_at"]) - entry["ended_at"]) <= 1
                            and int(row["loops"]) == entry["loops"]
                            and str(row["note"] or "").strip() == entry["note"] for row in rows)
                status, detail = (("duplicate", "同一段手动活动已存在") if exact else
                                  ("conflict", "同一时间已有不同的活动记录"))
        analysed.append({**entry, "status": status, "detail": detail,
                         "summary": _entry_summary(entry)})
    return analysed


def _entry_summary(entry: dict) -> str:
    when = _display_time(entry["occurred_at"])
    if entry["kind"] == "transaction":
        return f"{when} {entry['resource']} {entry['amount']:+d}"
    if entry["kind"] == "inventory":
        return f"{when} {entry['resource']} {entry['amount']}"
    return f"{when} {_SCRIPT_LABELS[entry['script']]} {entry['loops']} 圈"


def create_import_preview(store: TelemetryStore, data: bytes, filename: str) -> dict:
    entries, issues = _parse_rows(data, filename)
    analysed = _analyse_entries(store, entries)
    token = secrets.token_urlsafe(18)
    source_hash = hashlib.sha256(data).hexdigest()
    now = time.time()
    with _PREVIEW_LOCK:
        expired = [key for key, value in _PREVIEWS.items()
                   if now - value["created_at"] > PREVIEW_TTL_SECONDS]
        for key in expired:
            _PREVIEWS.pop(key, None)
        _PREVIEWS[token] = {"created_at": now, "entries": entries,
                            "filename": Path(filename).name, "sha256": source_hash}
    counts = {key: sum(item["status"] == key for item in analysed)
              for key in ("new", "duplicate", "conflict")}
    return {"schema_version": TRANSFER_SCHEMA_VERSION, "preview_id": token,
            "filename": Path(filename).name, "source_sha256": source_hash,
            "counts": {**counts, "invalid": sum(not item.get("ignored") for item in issues),
                       "ignored": sum(bool(item.get("ignored")) for item in issues)},
            "items": analysed, "issues": issues}


def _backup_database(store: TelemetryStore, backup_root: Path, metadata: dict) -> Path:
    stamp = datetime.now(_LEDGER_TZ).strftime("%Y%m%d-%H%M%S")
    folder = Path(backup_root) / f"ledger-import-{stamp}"
    suffix = 1
    while folder.exists():
        folder = Path(backup_root) / f"ledger-import-{stamp}-{suffix}"
        suffix += 1
    folder.mkdir(parents=True, exist_ok=False)
    target = folder / "telemetry.db"
    backup_conn = sqlite3.connect(str(target))
    try:
        store._conn().backup(backup_conn)
    finally:
        backup_conn.close()
    (folder / "manifest.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return folder


def apply_import_preview(store: TelemetryStore, preview_id: str, backup_root: Path,
                         *, accept_conflicts: bool = False) -> dict:
    with _PREVIEW_LOCK:
        preview = _PREVIEWS.get(str(preview_id))
    if not preview or time.time() - preview["created_at"] > PREVIEW_TTL_SECONDS:
        raise ValueError("这份预览已过期，请重新选文件")
    analysed = _analyse_entries(store, preview["entries"])
    conflicts = [item for item in analysed if item["status"] == "conflict"]
    if conflicts and not accept_conflicts:
        raise ValueError("预览后账本又发生了变化，或尚未确认冲突")
    selected = [item for item in analysed
                if item["status"] == "new" or (accept_conflicts and item["status"] == "conflict")]
    if not selected:
        return {"schema_version": TRANSFER_SCHEMA_VERSION, "imported": 0,
                "duplicates": sum(item["status"] == "duplicate" for item in analysed),
                "conflicts": len(conflicts), "backup": None}
    backup = _backup_database(store, Path(backup_root), {
        "kind": "ledger_import", "created_at": _display_time(time.time()),
        "source_filename": preview["filename"], "source_sha256": preview["sha256"],
        "planned_rows": len(selected),
    })
    conn = store._conn()
    imported = 0
    try:
        inventory_groups: dict[float, dict[str, int]] = defaultdict(dict)
        transaction_group_ids: dict[str, str] = {}
        for entry in selected:
            now = time.time()
            if entry["kind"] == "transaction":
                source_group = entry.get("group_key") or ""
                group_id = transaction_group_ids.setdefault(
                    source_group, secrets.token_hex(16)) if source_group else None
                conn.execute(
                    "INSERT INTO human_reports(created_at, occurred_at, source, gap_key, "
                    "activities, note, resource, claimed_delta, group_id) "
                    "VALUES (?, ?, 'proactive', NULL, ?, ?, ?, ?, ?)",
                    (now, entry["occurred_at"], _json(entry["activities"]), entry["note"],
                     entry["resource"], entry["amount"], group_id))
            elif entry["kind"] == "inventory":
                inventory_groups[entry["occurred_at"]][entry["resource"]] = entry["amount"]
            else:
                conn.execute(
                    "INSERT INTO manual_sessions(created_at, script, started_at, ended_at, loops, note) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (now, entry["script"], entry["occurred_at"], entry["ended_at"],
                     entry["loops"], entry["note"]))
            imported += 1
        for observed_at, resources in inventory_groups.items():
            payload = {"captured_at": _display_time(observed_at),
                       "source": "manual_import", "resources": resources}
            conn.execute(
                "INSERT INTO events(ts, run_id, script, event_type, payload) "
                "VALUES (?, NULL, 'manual', 'inventory.captured', ?)",
                (observed_at, _json(payload)))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    with _PREVIEW_LOCK:
        _PREVIEWS.pop(str(preview_id), None)
    return {"schema_version": TRANSFER_SCHEMA_VERSION, "imported": imported,
            "duplicates": sum(item["status"] == "duplicate" for item in analysed),
            "conflicts": len(conflicts), "backup": str(backup)}


def _export_snapshot(store: TelemetryStore) -> dict:
    conn = store._conn()
    candidates = []
    for sql in (
        "SELECT MIN(ts) FROM events", "SELECT MIN(occurred_at) FROM human_reports",
        "SELECT MIN(started_at) FROM manual_sessions",
    ):
        value = conn.execute(sql).fetchone()[0]
        if isinstance(value, (int, float)):
            candidates.append(float(value))
    now = time.time()
    start = min(candidates) if candidates else now - 86400
    ledger = store.resource_ledger(start, now)
    reports = conn.execute(
        "SELECT id, occurred_at, source, gap_key, activities, note, resource, "
        "claimed_delta, group_id FROM human_reports ORDER BY occurred_at, id").fetchall()
    inventories = conn.execute(
        "SELECT id, ts, payload FROM events WHERE run_id IS NULL AND script = 'manual' "
        "AND event_type = 'inventory.captured' ORDER BY ts, id").fetchall()
    sessions = conn.execute(
        "SELECT id, script, started_at, ended_at, loops, note FROM manual_sessions "
        "ORDER BY started_at, id").fetchall()
    return {"generated_at": now, "ledger": ledger,
            "reports": [{**dict(row), "activities": _loads(row["activities"], [])}
                        for row in reports],
            "inventories": [{"id": row["id"], "ts": row["ts"],
                              **_loads(row["payload"], {})} for row in inventories],
            "sessions": [dict(row) for row in sessions]}


def _full_ledger_rows(snapshot: dict) -> list[list[Any]]:
    rows = []
    for item in snapshot["ledger"]["attributions"]:
        rows.append([_display_time(item["ts"]), "まあ丸自动", "已归因收支",
                     _SCRIPT_LABELS.get(item.get("script"), item.get("script") or ""),
                     item["resource"], item["delta"], item.get("label") or "",
                     _confidence_label(item.get("confidence")), item.get("run_id") or "",
                     f"event:{item.get('event_id')}"])
    for report in snapshot["reports"]:
        if report["source"] != "proactive" or report["resource"] not in _RESOURCE_SET \
                or report["claimed_delta"] in (None, 0):
            continue
        rows.append([_display_time(report["occurred_at"]), "你手动", "手动收支", "",
                     report["resource"], report["claimed_delta"], report["note"], "手动",
                     "", f"manual:{report['id']}"])
    for gap in snapshot["ledger"]["gaps"]:
        if gap.get("human_report_ids"):
            continue
        for resource, delta in (gap.get("resources") or {}).items():
            rows.append([_display_time(gap["to"]), "未归因", "待确认差额", "", resource,
                         delta, gap.get("reason") or "", "低", "", gap["id"]])
    return sorted(rows, key=lambda row: (row[0], row[9]))


def _daily_rows(snapshot: dict) -> list[list[Any]]:
    manual = defaultdict(int)
    for report in snapshot["reports"]:
        if report["source"] == "proactive" and report["resource"] in _RESOURCE_SET \
                and isinstance(report["claimed_delta"], (int, float)):
            day = datetime.fromtimestamp(report["occurred_at"], _LEDGER_TZ).date().isoformat()
            manual[(day, report["resource"])] += int(report["claimed_delta"])
    daily = {(item["date"], item["resource"]): item for item in snapshot["ledger"]["daily_series"]}
    keys = sorted(set(daily) | set(manual))
    rows = []
    for key in keys:
        item = daily.get(key, {})
        rows.append([key[0], key[1], item.get("opening"), item.get("closing"),
                     item.get("total_delta"), item.get("attributed_delta", 0), manual.get(key, 0),
                     item.get("unattributed_delta"), item.get("observation_count", 0),
                     _confidence_label(item.get("confidence", "low"))])
    return rows


def _roundtrip_rows(snapshot: dict) -> list[list[Any]]:
    rows = []
    group_labels: dict[str, str] = {}
    for report in snapshot["reports"]:
        if report["source"] != "proactive" or report["resource"] not in _RESOURCE_SET \
                or report["claimed_delta"] in (None, 0):
            continue
        group_id = report.get("group_id") or ""
        group_label = group_labels.setdefault(
            group_id, f"手账组{len(group_labels) + 1}") if group_id else ""
        rows.append(["收支", _display_time(report["occurred_at"]), "", "", "",
                     report["resource"], report["claimed_delta"], report["note"],
                     ",".join(report["activities"]), group_label, "你手动"])
    for inventory in snapshot["inventories"]:
        if inventory.get("source") not in {"manual_entry", "manual_import"}:
            continue
        for resource, amount in (inventory.get("resources") or {}).items():
            if resource in _RESOURCE_SET and isinstance(amount, (int, float)):
                rows.append(["家底", _display_time(inventory["ts"]), "", "", "",
                             resource, amount, "", "", "", "你手动"])
    for session in snapshot["sessions"]:
        rows.append(["活动", _display_time(session["started_at"]),
                     _display_time(session["ended_at"]),
                     _SCRIPT_LABELS.get(session["script"], session["script"]), session["loops"],
                     "", 1, session["note"], "", "", "你手动"])
    return sorted(rows, key=lambda row: (row[1], row[0], row[5]))


def _add_table(sheet, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([_safe_cell(value) for value in row])
    if rows:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                               showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="315875")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    thin = Side(style="thin", color="D6D0C4")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    widths = []
    for column in range(1, len(headers) + 1):
        longest = max([len(str(headers[column - 1]))] +
                      [len(str(sheet.cell(row, column).value or "")) for row in range(2, min(sheet.max_row, 100) + 1)])
        widths.append(min(max(longest + 2, 10), 30))
        sheet.column_dimensions[get_column_letter(column)].width = widths[-1]


def export_ledger_xlsx(store: TelemetryStore) -> bytes:
    snapshot = _export_snapshot(store)
    workbook = Workbook()
    guide = workbook.active
    guide.title = "使用说明"
    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:F1")
    guide["A1"] = "まあ丸·本丸账房"
    guide["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="315875")
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 36
    guide["A3"] = "导出时间"
    guide["B3"] = _display_time(snapshot["generated_at"])
    guide["A5"] = "怎么看"
    guide["B5"] = "“完整流水”是まあ丸和你的收支明细；“当前家底”取最近一次可用观察；“每日汇总”按上海时间分日。"
    guide["A6"] = "怎么导回"
    guide["B6"] = "只编辑“可再次导入”。导入前会先预览、标出重复/冲突，真正写入前自动备份。"
    guide["A7"] = "数据边界"
    guide["B7"] = "まあ丸自动流水只能导出，不会被 Excel 反向改写。"
    for row in range(3, 8):
        guide[f"A{row}"].font = Font(name="Microsoft YaHei", bold=True, color="315875")
        guide[f"B{row}"].font = Font(name="Microsoft YaHei")
        guide[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 16
    guide.column_dimensions["B"].width = 82
    for row in (5, 6, 7):
        guide.row_dimensions[row].height = 38

    full = workbook.create_sheet("完整流水")
    _add_table(full, ["时间", "来源", "类型", "玩法", "资源", "变化", "说明", "可信度", "任务", "数据编号"],
               _full_ledger_rows(snapshot), "FullLedgerTable")

    holdings = workbook.create_sheet("当前家底")
    holding_rows = [[item["resource"], item["closing"], _confidence_label(item["confidence"]),
                     item["observation_count"], _display_time(snapshot["generated_at"])]
                    for item in snapshot["ledger"]["per_resource"]]
    _add_table(holdings, ["资源", "当前数量", "可信度", "窗口内观察数", "导出时间"],
               holding_rows, "CurrentHoldingsTable")

    daily = workbook.create_sheet("每日汇总")
    _add_table(daily, ["日期", "资源", "开始家底", "结束家底", "账面变化", "まあ丸已归因", "你手动记账", "未归因", "观察数", "可信度"],
               _daily_rows(snapshot), "DailySummaryTable")

    roundtrip = workbook.create_sheet(_IMPORT_SHEET)
    _add_table(roundtrip, ["记录类型", "时间", "结束时间", "玩法", "圈数", "资源", "数额", "备注", "标签", "组号", "来源"],
               _roundtrip_rows(snapshot), "RoundtripTable")
    roundtrip.sheet_properties.tabColor = "C48B35"

    for sheet in (full, holdings, daily, roundtrip):
        for column in (6, 7, 8):
            if column <= sheet.max_column:
                sheet.column_dimensions[get_column_letter(column)].width = min(
                    max(sheet.column_dimensions[get_column_letter(column)].width or 10, 13), 34)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_ledger_csv(store: TelemetryStore) -> bytes:
    snapshot = _export_snapshot(store)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["时间", "来源", "类型", "玩法", "资源", "变化", "说明", "可信度", "任务", "数据编号"])
    writer.writerows([_safe_cell(value) for value in row] for row in _full_ledger_rows(snapshot))
    return output.getvalue().encode("utf-8-sig")
